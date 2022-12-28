from openpyxl.cell.read_only import EmptyCell
import html


def getStartColumn(worksheet, cell_content):
    for current_row in worksheet.iter_rows(min_row=0, max_row=worksheet.max_row):
        # check EmptyRow and EmptyCell
        if current_row is not None and len(current_row) > 0:
            num_col = 0
            for cell in current_row:
                if cell is not None and type(cell) is not EmptyCell:
                    if cell_content in str(cell.value):
                        return num_col
                num_col += 1
    return -1

def getStartRow(worksheet, start_col, start_row):
    num_row = 0
    for current_row in worksheet.iter_rows(min_row=0, max_row=worksheet.max_row):
        # check EmptyRow and EmptyCell
        if current_row is not None and len(current_row) > 0:
            num_col = 0
            for cell in current_row:
                if num_col >= start_col and num_row >= start_row:
                    if cell is not None and type(cell) is not EmptyCell:
                        if 'Posición inicial' in str(cell.value):
                            return num_row
                num_col += 1
        num_row += 1
    return -1


def processMetadata(worksheet, start_col, start_row):
    metadata = []
    num_row = 0
    for current_row in worksheet.iter_rows(min_row=0, max_row=worksheet.max_row):
        cell = current_row[start_col]
        if cell is not None and type(cell) is not EmptyCell and len(str(cell.value)) > 0 and num_row > start_row:
            metadataRow = processMetadataRow(current_row, start_col)

            if metadataRow is not None:
                metadata.append(metadataRow)
            else:
                break
        num_row += 1
    return metadata



def getType(type_name):
    if type_name == 'Num':
        return 'NUMERIC'
    elif type_name == 'Char':
        return 'CHAR'
    return 'CHAR'

def getCanonicalName(file_name):
    offset = file_name.find('_') + 1
    if offset > 2:
        offset = 0
    file_name = file_name[offset:]
    file_name = file_name[:-4]
    return file_name


def getRegSize(metadata):
    size = 0
    for metadata_item in metadata:
        size += metadata_item['var_pos_inicial']
    return size

def updateOffset(metadata, global_offset):
    for metadata_item in metadata:
        metadata_item['var_pos_inicial'] = metadata_item['var_pos_inicial'] + global_offset

    return metadata

def getStartPoint(worksheet, cell_content):
    num_row = 0
    for current_row in worksheet.iter_rows(min_row=0, max_row=worksheet.max_row):
        if current_row is not None and len(current_row) > 0:
            num_col = 0
            for cell in current_row:
                if cell is not None and type(cell) is not EmptyCell:
                    if cell_content in str(cell.value):
                        return num_col, num_row
                num_col += 1
        num_row += 1
    return -1, -1

def getStartColRow(worksheet, file_name):
    start_col, start_row = getStartPoint(worksheet, file_name)
    if start_col < 0:
        print('ERROR reading header (start_col): ' + file_name)
    start_row = getStartRow(worksheet, start_col, start_row)
    if start_row < 0:
        print('ERROR reading header (start_col): ' + file_name)

    print(' start col:' + str(start_col))
    print(' start row:' + str(start_row))

    return start_col, start_row



def processMetadataRow(current_row, start_col):
    metadataMap = {}

    var_name = current_row[start_col+1]
    metadataMap['var_name'] = var_name.value

    var_pos_inicial = current_row[start_col]
    metadataMap['var_pos_inicial'] = var_pos_inicial.value

    var_tipo = current_row[start_col + 2]
    metadataMap['var_tipo'] = var_tipo.value

    var_long = current_row[start_col + 3]
    metadataMap['var_long'] = var_long.value

    var_decimales = current_row[start_col + 4]
    metadataMap['var_decimales'] = var_decimales.value

    var_desc = current_row[start_col + 5]
    metadataMap['var_desc'] = var_desc.value

    if var_tipo.value is None or var_name.value is None or var_long.value is None:
        return None

    return metadataMap



def getVarComunes(metadata, annios):
    var_comunes = {}
    # Contabiliza apariciones de cada variable
    for var in metadata:
        key = (var['var_name'], var['var_desc'])
        if key in var_comunes:
            var_comunes[key]['contador']+=1
        else:
            var_comunes[key] = var
            var_comunes[key]['contador']=1

    # Imprime las que no son comunes a los n años
    print("VARIABLES QUE NO APARECEN LOS N AÑOS:")
    for key in var_comunes:
        if var_comunes[key]['contador'] != len(annios):
            print(key,var_comunes[key]['contador'])
    print("\n\n")

    # Genera consulta SQL para consolidar en una unica tabla los campos comunes a las anuales.
    fields=[]
    for key in var_comunes:
        if var_comunes[key]['contador'] == len(annios):
            fields.append(var_comunes[key]['var_name'])

    return fields


def getRegSize(metadata):
    size = 0
    for metadata_item in metadata:
        size += metadata_item['var_pos_inicial']
    return size


def normalizeName(name, limit):
    normalized_name = ''
    normalized_name = name.replace('\n', ' ').replace('\r', '')
    normalized_name = html.escape(normalized_name)
    return normalized_name[:limit] + '..' * (len(normalized_name) > limit)