import openpyxl as xl
from openpyxl.cell.read_only import EmptyCell



iden_file_names  = ['1_IDEN2016.txt', '1_IDEN2017.txt', '1_IDEN2018.txt', '1_IDEN2019.txt']
renta_file_names = ['2_Renta2016.txt', '2_Renta2017.txt', '2_Renta2018.txt', '2_Renta2019.txt']
renta_imputacion_file_names = ['3_RentaImputacion2016.txt', '3_RentaImputacion2017.txt', '3_RentaImputacion2018.txt', '3_RentaImputacion2019.txt']
patrimonio_file_names = ['5_Patrimonio2016.txt', '5_Patrimonio2017.txt', '5_Patrimonio2018.txt', '5_Patrimonio2019.txt']
mod190_file_names = ['7_M190_2016.txt', '7_M190_2017.txt', '7_M190_2018.txt', '7_M190_2019.txt']

inputExcel = './doc/00_DiseñoRegistro_v3.xlsx'
output_folder = './out_unificado/'


def getStartColumn(worksheet, cell_content):
    for current_row in worksheet.iter_rows(min_row=0, max_row=worksheet.max_row):
        # check EmptyRow and EmptyCell
        if current_row is not None and len(current_row) > 0:
            num_col = 0
            for cell in current_row:
                if cell is not None and type(cell) is not EmptyCell:
                    if cell_content in str(cell.value):
                        return num_col
                num_col += 1
    return -1

def getStartRow(worksheet, start_col):
    num_row = 0
    for current_row in worksheet.iter_rows(min_row=0, max_row=worksheet.max_row):
        # check EmptyRow and EmptyCell
        if current_row is not None and len(current_row) > 0:
            num_col = 0
            for cell in current_row:
                if num_col >= start_col:
                    if cell is not None and type(cell) is not EmptyCell:
                        if 'Posición inicial' in str(cell.value):
                            return num_row
                num_col += 1
        num_row += 1
    return -1


def processMetadataRow(current_row, start_col, prefix):
    #{'var_name': 'IDENPER', 'var_pos_inicial': 1, 'var_tipo': 'Num', 'var_long': 11, 'var_decimales': None, 'var_desc': 'Identificador único de persona'}

    metadataMap = {}

    var_name = current_row[start_col + 1]
    metadataMap['var_name'] = var_name.value

    var_pos_inicial = current_row[start_col]
    metadataMap['var_pos_inicial'] = var_pos_inicial.value

    var_tipo = current_row[start_col + 2]
    metadataMap['var_tipo'] = var_tipo.value

    var_long = current_row[start_col + 3]
    metadataMap['var_long'] = var_long.value

    var_decimales = current_row[start_col + 4]
    metadataMap['var_decimales'] = var_decimales.value

    var_desc = current_row[start_col + 5]
    metadataMap['var_desc'] = var_desc.value

    if var_tipo.value is None or var_name.value is None or var_tipo.value is None or var_long.value is None:
        return None

    metadataMap['var_name'] = prefix + '_' + metadataMap['var_name']

    return metadataMap


def processMetadata(worksheet, start_col, start_row, prefix):
    metadata = []
    num_row = 0
    for current_row in worksheet.iter_rows(min_row=0, max_row=worksheet.max_row):
        cell = current_row[start_col]
        if cell is not None and type(cell) is not EmptyCell and len(str(cell.value)) > 0 and num_row > start_row:
            metadataRow = processMetadataRow(current_row, start_col, prefix)
            if metadataRow is not None:
                metadata.append(metadataRow)
        num_row += 1
    return metadata


def getType(type_name):
    if type_name == 'Num':
        return 'NUMERIC'
    elif type_name == 'Char':
        return 'CHAR'
    return 'CHAR'

def getCanonicalName(file_name):
    offset = file_name.find('_') + 1
    file_name = file_name[offset:]
    file_name = file_name[:-4]
    return file_name


def writeLoadData(metadata, original_file_name, annio):
    # file_name = getCanonicalName(original_file_name)
    # file_name_def = output_folder + 'load_' + file_name + '.sql'
    # tablename = 'tbl_' + file_name
    #
    # m = re.findall(r'[0-9]{4,7}', original_file_name)
    # annio = str(m[0])
    #
    # if original_file_name.startswith('IRPF20'):
    #     original_file_name = '10_' + original_file_name
    #
    # with open(file_name_def, 'w') as f:
    #     original_file_name = data_folder + annio + '/' + original_file_name
    #     f.write('USE ' + database_name + ';\n\n')
    #     f.write('LOAD DATA LOCAL INFILE \'' + original_file_name + '\'\n')
    #     f.write('INTO TABLE ' + tablename + ' FIELDS TERMINATED BY \'\';\n')



    #file_name = getCanonicalName(original_file_name)
    file_name_def = output_folder + str(annio) + '/load_' + original_file_name + '.sql'
    tablename = 'tbl_' + file_name

    with open(file_name_def, 'w') as f:
        f.write('USE test;\n\n')
        f.write('LOAD DATA LOCAL INFILE \'' + original_file_name + '\'\n')
        f.write('INTO TABLE ' + tablename + '\n')
        f.write('(@row)\n')
        f.write('SET \n')

        for metadata_item in metadata:
            var_name = metadata_item.get('var_name')
            var_pos_inicial = metadata_item.get('var_pos_inicial')
            var_long = metadata_item.get('var_long')

            if metadata_item == metadata[-1]:
                f.write('\t' + var_name + ' = TRIM(SUBSTR(@row,' + str(var_pos_inicial) + ', ' + str(var_long) + '))\n')
            else:
                f.write('\t' + var_name + ' = TRIM(SUBSTR(@row,' + str(var_pos_inicial) + ', ' + str(var_long) + ')),\n')

        f.write(';')



def writeCreateTable(metadata, original_file_name, annio):
    #file_name = getCanonicalName(original_file_name)
    file_name_def = output_folder + str(annio) + '/create_table_' + original_file_name + '.sql'
    tablename = 'tbl_' + file_name

    with open(file_name_def, 'w') as f:
        f.write('USE test;\n\n')
        f.write('DROP TABLE IF EXISTS ' + tablename + ';\n\n')
        f.write('CREATE TABLE ' + tablename + '(\n')

        for metadata_item in metadata:
            var_name = metadata_item.get('var_name')
            var_tipo = getType(metadata_item.get('var_tipo'))
            var_long = metadata_item.get('var_long')
            var_decimales = metadata_item.get('var_decimales')

            strLength = ''
            if var_decimales is not None:
                strLength = str(var_long) + ','+ str(var_decimales)
            else:
                strLength = str(var_long)

            if metadata_item == metadata[-1]:
                f.write('\t' + var_name + ' ' + var_tipo + '(' + strLength + ') DEFAULT NULL\n')
                # TODO var_desc
            else:
                f.write('\t' + var_name + ' ' + var_tipo + '(' + strLength + ') DEFAULT NULL,\n')

        f.write(') engine=columnstore CHARSET=latin1 COLLATE=latin1_spanish_ci;')


def getStartColRow(worksheet, file_name):
    start_col = getStartColumn(worksheet, file_name)
    if start_col < 0:
        print('ERROR reading header (start_col): ' + file_name)
    start_row = getStartRow(worksheet, start_col)
    if start_row < 0:
        print('ERROR reading header (start_col): ' + file_name)

    print(' start col:' + str(start_col))
    print(' start row:' + str(start_row))

    return start_col, start_row


def getMetadata(file_name, worksheet, prefix):
    start_col, start_row = getStartColRow(worksheet, file_name)
    if start_col < 0 or start_row < 0:
        return None
    metadata = processMetadata(worksheet, start_col, start_row, prefix)
    return metadata

def getRegSize(metadata):
    size = 0
    for metadata_item in metadata:
        size += metadata_item['var_pos_inicial']
    return size

def updateOffset(metadata, global_offset):
    for metadata_item in metadata:
        metadata_item['var_pos_inicial'] = metadata_item['var_pos_inicial'] + global_offset

    return metadata

def getTableNameAndPrefix(title, num_annio):
    file_name = ''
    prefix = ''
    if worksheet.title == '1_IDEN':
        file_name = iden_file_names[num_annio]
        prefix = 'IDEN'
    elif worksheet.title == '2_Renta':
        file_name = renta_file_names[num_annio]
        prefix = 'RENTA'
    elif worksheet.title == '3_RentaImputación':
        file_name = renta_imputacion_file_names[num_annio]
        prefix = 'RENTA_IMPUT'
    elif worksheet.title == '5_Patrimonio':
        file_name = patrimonio_file_names[num_annio]
        prefix = 'PATRIM'
    elif worksheet.title == '7_Mod190':
        file_name = mod190_file_names[num_annio]
        prefix = 'M190'
    return file_name, prefix




# process worksheets
workbook = xl.load_workbook(inputExcel, keep_links=False, read_only=True, data_only=True)

metadata = []
annios = [2016, 2017, 2018, 2019]
#annios = [2016]
global_offset = 0
processed_file_names = []


def writePasteCmd(processed_file_names, annio):
    file_name = output_folder + str(annio) + '/paste_cmd' + str(annio) + '.sh'

    with open(file_name, 'w') as f:
        f.write('# dos2unix *.txt\n')
        f.write('# dos2unix *.TXT\n')
        f.write('paste -d"\\0" ')
        for processed_file_name in processed_file_names:
            f.write(' ' + processed_file_name)
        f.write(' > ' + str(annio) + '_unificado.txt')


var_comunes={}
for annio in annios:
    for worksheet in workbook.worksheets:
        num_annio = annio - 2016

        print('Sheet: ' + worksheet.title)
        file_name, prefix = getTableNameAndPrefix(worksheet.title, num_annio)
        if len(file_name) == 0:
            continue

        table_metadata = getMetadata(file_name, worksheet, prefix)
        if table_metadata is None:
            continue
        table_metadata = updateOffset(table_metadata, global_offset)
        metadata.extend(table_metadata)

        reg_size = getRegSize(table_metadata)
        global_offset += reg_size

        processed_file_names.append(file_name)

    file_name = 'unificado_' + str(annio)
    writeCreateTable(metadata, file_name, annio)
    writeLoadData(metadata, file_name, annio)
    writePasteCmd(processed_file_names, annio)

# ==========================================================
#    Generar tabla Unificada (select union de las anuales)
# ==========================================================

# Contabiliza apariciones de cada variable
for var in metadata:
    key=(var['var_name'],var['var_desc'])
    if key in var_comunes:
        var_comunes[key]['contador']+=1
    else:
        var_comunes[key]=var
        var_comunes[key]['contador']=1

# Imprime las que no son comunes a los n años
print("VARIABLES QUE NO APARECEN LOS N AÑOS:")
for key in var_comunes:
    if var_comunes[key]['contador']!=len(annios):
        print(key,var_comunes[key]['contador'])
print("\n\n")

# Genera consulta SQL para consolidar en una unica tabla los campos comunes a las anuales.
fields=[]
for key in var_comunes:
    if var_comunes[key]['contador']==len(annios):
        fields.append(var_comunes[key]['var_name'])

fields=",".join(sorted(fields))
tablas=[f"tbl_{a}_hogares" for a in annios]
tablas=" UNION ALL ".join([f"SELECT * FROM {t}" for t in tablas ])
sql=f"CREATE TABLE tbl_unificado_hogares AS SELECT {fields} FROM ({tablas});"

# Escribe SQL UNION ALL a fichero
with open("out_unificado/create_global_hogares.sql", 'w') as f:
     f.write(sql+"\n")


# ==========================================================
#                Generar Esquema de Mondrian)
# ==========================================================
colores={'M190':'#4469a6','PATRIM':'#0b4d90','RENTA':'#3399ff'}

measures=[]
dimensiones=[]
head="""
<?xml version="1.0" encoding="ISO-8859-1" standalone="yes"?>
<Schema name="hogares" measuresCaption="Variables de explotaci&#243;n">
<Cube name="anuarioTotal" visible="true" cache="true" enabled="true" caption="Anuario estad&#237;stico">
"""

foot="""\n\n	</Cube>
</Schema>"""

for var in metadata:
    key=(var['var_name'],var['var_desc'])
    if var_comunes[key]['contador']!=len(annios): continue

    v=var_comunes[key]
    tabla=v['var_name'].split("_")[0]

    if tabla=='IDEN':
        #print("EXCLUIDO:"+tabla)
        dimension=f'''
		<Dimension name="{tabla}" caption="{tabla}" description="{tabla}">
		  <Hierarchy hasAll="false" allMemberName="total">
			<Level caption="{tabla}" name="{v['var_name']}" column="{v['var_name']}" uniqueMembers="true"  captionColumn="columnName_es">
			</Level>
		  </Hierarchy>
		</Dimension>
        '''
        dimensiones.append(dimension)

    else:
        measure=f'''
        <Measure name="{v['var_name']}"  column="{v['var_name']}"  formatString="#,###;-#,###;0" aggregator="sum" caption="{var['var_name']}" description="{var['var_desc']}">
                <CalculatedMemberProperty name="DISPLAY_FOLDER" value ="{tabla}|color:{colores[tabla]}"/>
        </Measure>
        '''
        measures.append(measure)
        
esquema=head+"\n"+"\n".join(dimensiones)+"\n".join(measures)+foot
print(esquema)
print(dimension)
# Escribe SQL UNION ALL a fichero
with open("out_unificado/hogares.xml", 'w') as f:
     f.write(esquema+"\n")






